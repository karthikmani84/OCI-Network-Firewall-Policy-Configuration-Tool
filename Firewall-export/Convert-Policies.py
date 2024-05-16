import pandas as pd
from openpyxl import load_workbook
import json
import os

# Function to extract port ranges
def extract_port_ranges(port_ranges):
    minimum_ports = []
    maximum_ports = []
    for port_range in port_ranges:
        minimum_ports.append(port_range.get('minimum-port'))
        maximum_ports.append(port_range.get('maximum-port'))
    return minimum_ports, maximum_ports

# Function to load JSON data
def load_json(file_path):
    with open(file_path, "r") as file:
        return json.load(file)

# Load existing Excel file if it exists
output_file = "output.xlsx"
existing_sheets = {}
if os.path.exists(output_file):
    os.remove(output_file)
    pd.DataFrame().to_excel(output_file, index=False)
else:
    # Create the Excel file with an empty DataFrame
    pd.DataFrame().to_excel(output_file, index=False)

# Process 'security-rule' data
security_rules_data = load_json("security_rule_output.json")
security_rules_df = pd.DataFrame()
for rule in security_rules_data:
    rule_data = rule.get('data', {})
    condition_tree = rule_data.get('condition', {})
    security_rules_df = pd.concat([security_rules_df, pd.DataFrame({
        'name': [rule_data.get('name')],
        'Source Address Lists': [", ".join(condition_tree.get('source-address', []))],
        'Destination Address Lists': [", ".join(condition_tree.get('destination-address', []))],
        'Service Lists': [", ".join(condition_tree.get('service', []))],
        'Application Lists': [", ".join(condition_tree.get('application', []))],
        'Url Lists': [", ".join(condition_tree.get('url', []))],
        'Action': [rule_data.get('action')]
    })])

# Process 'url-list' data
url_list_data = load_json("url_list_output.json")
urls_to_append = []
url_list_df = pd.DataFrame()
for item in url_list_data:
    name = item['data'].get('name')
    #print(item['data'].get('name'))
    urls = item['data'].get('urls', [])
    #print(item['data'].get('urls', []))
    for url in urls:
        pattern = url.get('pattern', '')
        urls_to_append.append({'name': name, 'pattern': pattern})
        url_list_df = pd.DataFrame(urls_to_append)
#print(url_list_df)

# Process 'iplist' data
iplist_data = load_json("addresslist_output.json")
iplist_df = pd.DataFrame()
for entry in iplist_data:
    entry_data = entry.get('data', {})
    iplist_df = pd.concat([iplist_df, pd.DataFrame({
        'name': [entry_data.get('name')],
        'addresses': [", ".join(entry_data.get('addresses', []))]
    })])

# Write 'security-rules' and 'iplist' and url_lists sheets
with pd.ExcelWriter(output_file, mode='a', if_sheet_exists='replace') as writer:
    security_rules_df.to_excel(writer, sheet_name='security-rules', index=False)
    iplist_df.to_excel(writer, sheet_name='iplist', index=False)
    url_list_df.to_excel(writer, sheet_name='url_lists', index=False)


# Load existing 'service' sheet from the output Excel file or create it if it doesn't exist
if 'service' in existing_sheets:
    existing_services_df = existing_sheets['service']
else:
    existing_services_df = pd.DataFrame()

# Process 'service' data
service_data = load_json("service_output.json")
service_individual_df = pd.DataFrame()  # Separate dataframe for individual services
for entry in service_data:
    entry_data = entry.get('data', {})
    minimum_ports, maximum_ports = extract_port_ranges(entry_data.get('port-ranges', []))
    service_individual_df = pd.concat([service_individual_df, pd.DataFrame({
        'name': [entry_data.get('name')],
        'type': [entry_data.get('type')],
        'minimumPort': [", ".join(map(str, minimum_ports))],
        'maximumPort': [", ".join(map(str, maximum_ports))]
    })])

# Process 'Application list' data
Application_group_data = load_json("application_output.json")
Application_group_df = pd.DataFrame()  # Separate dataframe for application groups
# Initialize lists to store data
app_names = []
icmp_types = []

# Extract data from JSON
for entry in Application_group_data:
    data = entry.get('data', {})
    app_names.append(data.get('name'))
    icmp_types.append(data.get('icmp-type'))

# Create DataFrame from new data
df_Application = pd.DataFrame({
    'name': app_names,
    'type': 'ICMP_TYPE',
    'icmpType': icmp_types
})

# Process 'service list' data
service_group_data = load_json("servicelist_output.json")
service_group_df = pd.DataFrame()  # Separate dataframe for service groups
# Initialize lists to store data
group_names = []
group_services = []

# Extract data from JSON
for entry in service_group_data:
    data = entry.get('data', {})
    services_list = data.get('services', [])
    if len(services_list) > 1:
        group_names.append(data.get('name'))
        group_services.append(", ".join(services_list))

# Create DataFrame from new data
df_new_services = pd.DataFrame({
    'name': group_names,
    'type': 'SERVICE_GROUP',
    'services': group_services
})

# Process 'Application list' data
Application_group_data = load_json("applicationlist_output.json")
Application_group_df = pd.DataFrame()  # Separate dataframe for service groups
# Initialize lists to store data
grp_names = []
grp_services = []

# Extract data from JSON
for entry in Application_group_data:
    data = entry.get('data', {})
    services_list = data.get('apps', [])
    if len(services_list) > 1:
        grp_names.append(data.get('name'))
        grp_services.append(", ".join(services_list))

# Create DataFrame from new data
df_applist_services = pd.DataFrame({
    'name': grp_names,
    'type': 'ICMP_GROUP',
    'services': grp_services
})

# Write 'service' sheet
with pd.ExcelWriter(output_file, mode='a', if_sheet_exists='replace') as writer:
    combined_services_df = pd.concat([existing_services_df, service_individual_df, df_new_services, df_Application, df_applist_services], ignore_index=True)
    combined_services_df.to_excel(writer, sheet_name='service', index=False)
    # Write other sheets
    for sheet_name, df in existing_sheets.items():
        if sheet_name not in ['security-rules', 'iplist', 'service', 'url_lists']:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
print()

#makes sure sheet1 isnt in the file.
wb = load_workbook("output.xlsx")
# Check if the sheet exists
if "Sheet1" in wb.sheetnames:
    # Get the sheet
    sheet = wb["Sheet1"]
    # Remove the sheet
    wb.remove(sheet)

# Save the changes
wb.save("output.xlsx")

print("Excel file created and updated successfully.")
